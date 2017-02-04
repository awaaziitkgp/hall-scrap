from bs4 import BeautifulSoup
import openpyxl
wb = openpyxl.load_workbook('cg.xlsx')
f = open('hmc.txt', 'r')

soup = BeautifulSoup(f.read(), "html.parser")

sheet = wb.get_sheet_by_name('Sheet1')
n=sheet.max_row+1

for i in range(1,n):
	roll= sheet.cell(row=i, column=1).value
	for tag in soup.find_all(text=roll):
		hall=tag.parent.next_sibling.next_sibling.string
	if soup.find_all(text=roll):
		sheet['B'+str(i)]=hall

wb.save('cg.xlsx')

